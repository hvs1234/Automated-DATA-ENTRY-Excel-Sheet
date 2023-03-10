from tkinter import * 
from tkinter import messagebox,colorchooser
from tkinter.ttk import Combobox
import openpyxl ; from openpyxl import Workbook 
import xlrd
import pathlib

#Application Setup
root = Tk()
root.title("Data Entry")
root.resizable(False,False)
root.geometry("700x420+340+180")
root.configure(bg="#326273")

#Functions
file = pathlib.Path('BackEnd_Data.xlsx')
if(file.exists()): pass
else: 
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone No."
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save("BackEnd_Data.xlsx")

def color(event):
    cls = colorchooser.askcolor(title="Select Color To Change")
    root.configure(bg=cls[1])
    l1.config(bg=cls[1])
    l2.config(bg=cls[1])
    l3.config(bg=cls[1])
    l4.config(bg=cls[1])
    l5.config(bg=cls[1])
    l6.config(bg=cls[1])
    b1.config(bg=cls[1],activebackground=cls[1],activeforeground="#fff")
    b2.config(bg=cls[1],activebackground=cls[1],activeforeground="#fff")
    b3.config(bg=cls[1],activebackground=cls[1],activeforeground="#fff")

def submit():
    name = name_value.get()
    contact = contact_value.get()
    age = age_value.get()
    gender = gender_box.get()
    address = address_entry.get(1.0,END)

    if(name=='' and contact=='' and age==''):
        messagebox.showwarning("Blank Input","Please fill input this field!")

    else:
        file = openpyxl.load_workbook("BackEnd_Data.xlsx")
        sheet = file.active
        sheet.cell(column = 1 , row = sheet.max_row+1 , value=name)
        sheet.cell(column = 2 , row = sheet.max_row , value=contact)
        sheet.cell(column = 3 , row = sheet.max_row , value=age)
        sheet.cell(column = 4 , row = sheet.max_row , value=gender)
        sheet.cell(column = 5 , row = sheet.max_row , value=address)
        file.save(r'BackEnd_Data.xlsx')

def delete():
    name_value.set('')
    contact_value.set('')
    age_value.set('')
    address_entry.delete(1.0,END)

#Application Creation
img1 = PhotoImage(file="E:\\pyImages\\data entry logo.png")
root.iconphoto(False,img1)
l1 = Label(root,text="Please Fill Out This Entry Form!",font="Gabriola 18 bold",bg="#326273",fg="#fff")
l1.place(x=20,y=20)

l2 = Label(root,text="Name",font="Algerian 15",bg="#326273",fg="#fff")
l2.place(x=50,y=100)
########################################################################
l3 = Label(root,text="Contact No.",font="Algerian 15",bg="#326273",fg="#fff")
l3.place(x=50,y=150)
########################################################################
l4 = Label(root,text="Age",font="Algerian 15",bg="#326273",fg="#fff")
l4.place(x=50,y=200)
########################################################################
l5 = Label(root,text="Gender",font="Algerian 15",bg="#326273",fg="#fff")
l5.place(x=365,y=200)
########################################################################
l6 = Label(root,text="Address",font="Algerian 15",bg="#326273",fg="#fff")
l6.place(x=50,y=250)
########################################################################
########################################################################
name_value = StringVar()
contact_value = StringVar()
age_value = StringVar()
########################################################################
name_entry = Entry(root,textvariable=name_value,width=36,bd=4,font=("Lucida Calligraphy" ,10,"bold"),
bg="light grey",fg="indigo")
name_entry.place(x=200,y=100)
########################################################################
contact_entry = Entry(root,textvariable=contact_value,width=36,bd=4,font=("Lucida Calligraphy" ,10,"bold"),
bg="light grey",fg="indigo")
contact_entry.place(x=200,y=150)
########################################################################
age_entry = Entry(root,textvariable=age_value,width=14,bd=4,font=("Lucida Calligraphy" ,10,"bold"),
bg="light grey",fg="indigo")
age_entry.place(x=200,y=200)
########################################################################
gender_box = Combobox(root,values=['Male','Female','Transgender'],
font="Arial 10 bold",state="r",width=12)
gender_box.place(x=460,y=200) ; gender_box.set('--Select--')
########################################################################
address_entry = Text(root,width=36,height=4,bd=4,font=("Lucida Calligraphy" ,10,"bold"),
bg="light grey",fg="indigo",wrap=WORD)
address_entry.place(x=200,y=250)
########################################################################
b1 = Button(root,text="Submit",bg="#326273",fg="#fff",width=16,bd=4,height=2,
activebackground="#326273",activeforeground="#fff",command=submit)
b1.place(x=200,y=350)
b2 = Button(root,text="Clear",bg="#326273",fg="#fff",width=16,bd=4,height=2,
activebackground="#326273",activeforeground="#fff",command=delete)
b2.place(x=340,y=350)
b3 = Button(root,text="Exit",bg="#326273",fg="#fff",width=12,bd=4,height=2,
activebackground="#326273",activeforeground="#fff",command=root.destroy)
b3.place(x=480,y=350)

root.bind('<Control-g>',color)
root.mainloop()

